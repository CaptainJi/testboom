import json
from pathlib import Path
from typing import Dict, Any, Optional, List
import pandas as pd
from ..ai_core.chat_manager import ChatManager
from ..logger.logger import logger
from .file_processor import FileProcessor
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

class DocAnalyzer:
    """文档分析器
    
    负责分析PRD文档，包括：
    1. 文档解压和处理
    2. 图片内容分析
    3. 需求理解和结构化
    4. 测试用例生成
    5. 结果导出
    """
    
    def __init__(self, work_dir: Optional[str] = None) -> None:
        """初始化文档分析器
        
        Args:
            work_dir: 工作目录,默认为当前目录下的temp
        """
        self.work_dir = Path(work_dir) if work_dir else Path.cwd() / 'temp'
        self.file_processor = FileProcessor(str(self.work_dir))
        self.chat_manager = ChatManager()
    
    def analyze_prd(self, zip_path: str) -> Optional[Dict[str, Any]]:
        """分析PRD文档
        
        Args:
            zip_path: PRD文档zip文件路径
            
        Returns:
            Optional[Dict[str, Any]]: 分析结果
        """
        try:
            # 解压文件
            files = self.file_processor.extract_zip(zip_path)
            if not files:
                logger.error(f"解压文件失败或文件为空: {zip_path}")
                return None
            
            # 分类文件
            classified = self.file_processor.classify_files(files)
            logger.info(f"文件分类结果: {classified}")
            
            # 分析图片
            image_results = []
            for image_path in classified['images']:
                logger.info(f"\n开始分析图片: {image_path}")
                
                # 验证图片
                if not self.file_processor.validate_image(image_path):
                    logger.warning(f"跳过无效图片: {image_path}")
                    continue
                
                # 分析图片
                result = self.chat_manager.chat_with_images(
                    "请分析这张图片，提取所有对测试用例设计有帮助的信息。",
                    [str(image_path)]
                )
                
                if result:
                    logger.info(f"\n图片分析结果:\n{result}")
                    image_results.append({
                        'file': str(image_path),
                        'content': result,
                        'features': self._extract_features(result)
                    })
                else:
                    logger.error(f"分析图片失败: {image_path}")
            
            # 构建汇总内容
            logger.info("\n开始构建汇总内容...")
            summary_content = self._build_summary_content(image_results)
            
            # 生成汇总报告
            logger.info("\n开始生成汇总报告...")
            summary = self.chat_manager.analyze_requirement(summary_content)
            if not summary:
                logger.error("生成汇总报告失败")
                return None
            
            # 生成测试用例
            logger.info("\n开始生成测试用例...")
            testcases = self.chat_manager.generate_testcases(
                summary,
                {'images': image_results}
            )
            if not testcases:
                logger.error("生成测试用例失败")
                return None
            
            # 导出测试用例到Excel
            if testcases:
                output_path = self.work_dir / 'testcases.xlsx'
                self._export_testcases_to_excel(testcases, str(output_path))
            
            # 清理临时文件
            self.file_processor.cleanup()
            
            return {
                'summary': summary,
                'testcases': testcases,
                'details': {
                    'images': image_results
                }
            }
            
        except Exception as e:
            logger.error(f"分析PRD文档失败: {e}")
            logger.exception(e)
            return None
    
    def _build_summary_content(self, image_results: List[Dict[str, Any]]) -> str:
        """构建汇总内容
        
        Args:
            image_results: 图片分析结果列表
            
        Returns:
            str: 汇总内容
        """
        summary_content = ""
        for result in image_results:
            summary_content += f"\n文件名: {Path(result['file']).name}\n"
            summary_content += f"类型: image\n"
            summary_content += f"分析结果:\n{result['content']}\n"
            summary_content += "=" * 50 + "\n"
        return summary_content
    
    def _extract_features(self, content: str) -> Dict[str, List[str]]:
        """从分析结果中提取特征
        
        Args:
            content: 分析结果文本
            
        Returns:
            Dict[str, List[str]]: 提取的特征
        """
        try:
            features = {
                'functionality': [],  # 功能点
                'workflow': [],  # 业务流程
                'data_flow': [],  # 数据流向
                'interfaces': [],  # 接口定义
                'constraints': [],  # 约束条件
                'exceptions': []  # 异常场景
            }
            
            # 按段落分割内容
            paragraphs = content.split('\n\n')
            
            current_section = None
            for para in paragraphs:
                para = para.strip()
                if not para:
                    continue
                
                # 识别段落类型
                if '功能点' in para or '功能列表' in para:
                    current_section = 'functionality'
                elif '业务流程' in para or '操作步骤' in para:
                    current_section = 'workflow'
                elif '数据流' in para:
                    current_section = 'data_flow'
                elif '接口' in para:
                    current_section = 'interfaces'
                elif '约束' in para or '限制' in para:
                    current_section = 'constraints'
                elif '异常' in para or '错误' in para:
                    current_section = 'exceptions'
                elif current_section:
                    # 将内容添加到当前段落类型
                    features[current_section].append(para)
            
            return features
            
        except Exception as e:
            logger.error(f"提取特征失败: {e}")
            logger.exception(e)
            return {k: [] for k in ['functionality', 'workflow', 'data_flow', 'interfaces', 'constraints', 'exceptions']}
    
    def _export_testcases_to_excel(self, testcases: List[Dict[str, Any]], output_path: str) -> None:
        """导出测试用例到Excel
        
        Args:
            testcases: 测试用例列表
            output_path: 输出文件路径
        """
        try:
            # 转换测试用例格式
            data = []
            for tc in testcases:
                # 为步骤和预期结果添加项目符号
                steps = [f"• {step}" for step in tc.get('steps', [])]
                expected = [f"• {exp}" for exp in tc.get('expected', [])]
                
                data.append({
                    '用例ID': tc.get('id', ''),
                    '所属模块': tc.get('module', ''),
                    '用例名称': tc.get('name', ''),
                    '用例等级': tc.get('level', ''),
                    '前置条件': tc.get('precondition', ''),
                    '测试步骤': '\n'.join(steps),
                    '预期结果': '\n'.join(expected),
                    '实际结果': '',
                    '测试状态': '未执行',
                    '备注': tc.get('notes', '')
                })
            
            # 创建DataFrame并导出
            df = pd.DataFrame(data)
            
            # 创建Excel writer对象
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # 写入数据
                df.to_excel(writer, index=False, sheet_name='测试用例')
                
                # 获取工作簿和工作表
                workbook = writer.book
                worksheet = writer.sheets['测试用例']
                
                # 定义样式
                header_fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
                header_font = Font(bold=True, size=11)
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell_alignment = Alignment(vertical='center', wrap_text=True)
                border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                # 设置列宽
                column_widths = {
                    'A': 12,  # 用例ID
                    'B': 15,  # 所属模块
                    'C': 20,  # 用例名称
                    'D': 10,  # 用例等级
                    'E': 25,  # 前置条件
                    'F': 40,  # 测试步骤
                    'G': 40,  # 预期结果
                    'H': 15,  # 实际结果
                    'I': 12,  # 测试状态
                    'J': 20,  # 备注
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # 应用表头样式
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border
                
                # 获取数据范围
                data_rows = len(df) + 1
                data_cols = len(df.columns)
                
                # 应用数据单元格样式
                for row in range(2, data_rows + 1):
                    # 设置行高
                    worksheet.row_dimensions[row].height = 30
                    
                    # 设置交替行背景色
                    row_fill = PatternFill(
                        start_color='F5F5F5' if row % 2 == 0 else 'FFFFFF',
                        end_color='F5F5F5' if row % 2 == 0 else 'FFFFFF',
                        fill_type='solid'
                    )
                    
                    for col in range(1, data_cols + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.alignment = cell_alignment
                        cell.border = border
                        cell.fill = row_fill
                
                # 冻结首行
                worksheet.freeze_panes = 'A2'
                
            logger.info(f"测试用例已导出到: {output_path}")
            
        except Exception as e:
            logger.error(f"导出测试用例失败: {e}")
            logger.exception(e)
